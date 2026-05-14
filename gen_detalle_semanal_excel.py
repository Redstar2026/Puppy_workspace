"""
Genera Detalle_Semanal_Todos_Paises.xlsx
Hojas:
  - RESUMEN        : totales consolidados por pais
  - CR/GT/HN/NI/SV : Detalle Semanal (Normal/Oferta/Mayoreo + deltas SW14vs13 y SW14vs12)
  - TENDENCIAS     : Clasificacion Creciente vs Decreciente con deltas y formato condicional
Formato condicional: verde (positivo), rojo (negativo), amarillo (cero)
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from datetime import date

# ──────────────────────────────────────────────────────────────────
#  DATOS  (tot = total registros, incluye nor+ofe+may solapados)
# ──────────────────────────────────────────────────────────────────
DATA = {
 "CR": [
  {"c":"AUTOMERCADO",        "tot":[9445,9247,9121,8955],    "nor":[9415,9247,9121,8939],    "ofe":[942,1101,1150,1172],    "may":[0,0,0,0],            "upcs":[1306,1278,1260,1238],"cats":[131,130,130,130]},
  {"c":"AVENIDA 10",         "tot":[8962,8905,9139,8855],    "nor":[8932,8897,9135,8855],    "ofe":[309,299,356,325],       "may":[117,86,165,192],     "upcs":[1249,1241,1274,1236],"cats":[114,112,113,112]},
  {"c":"BM",                 "tot":[3458,3599,3556,3664],    "nor":[3458,3493,3514,3661],    "ofe":[9,155,198,82],          "may":[0,11,30,0],          "upcs":[486,512,501,515],    "cats":[109,113,112,112]},
  {"c":"CENTRAL DE COMPRAS", "tot":[2401,2156,2289,2065],    "nor":[2401,2156,2289,2065],    "ofe":[6,0,25,0],              "may":[132,211,167,266],    "upcs":[339,305,323,292],    "cats":[86,81,85,79]},
  {"c":"CENTRO DE COMPRAS",  "tot":[7567,7372,7756,7322],    "nor":[7567,7371,7756,7322],    "ofe":[6,7,12,6],              "may":[222,230,168,144],    "upcs":[558,551,578,549],    "cats":[95,95,96,90]},
  {"c":"COMPREBIEN",         "tot":[5418,5411,5411,5411],    "nor":[5418,5411,5411,5411],    "ofe":[138,162,162,156],       "may":[0,0,0,0],            "upcs":[759,758,758,758],    "cats":[99,99,99,99]},
  {"c":"ECONOMAS",           "tot":[3360,3279,3145,2804],    "nor":[3360,3213,3115,2744],    "ofe":[0,108,186,120],         "may":[0,0,0,0],            "upcs":[474,464,442,400],    "cats":[105,105,100,100]},
  {"c":"ECONOMICO",          "tot":[4438,4391,4398,3807],    "nor":[4438,4347,4368,3801],    "ofe":[12,66,48,14],           "may":[9,34,12,17],         "upcs":[412,410,400,368],    "cats":[98,98,99,97]},
  {"c":"EL EMPINO",          "tot":[4872,4858,4976,4966],    "nor":[4872,4858,4942,4942],    "ofe":[12,0,44,24],            "may":[0,46,0,10],          "upcs":[343,341,349,349],    "cats":[91,91,91,94]},
  {"c":"FARMA VALUE",        "tot":[8295,8204,8197,8211],    "nor":[8295,8204,8197,8211],    "ofe":[0,0,0,0],               "may":[0,0,0,0],            "upcs":[1181,1168,1167,1169],"cats":[21,21,21,21]},
  {"c":"FARMACIA LA BOMBA",  "tot":[7406,7315,7322,7301],    "nor":[7406,7315,7322,7301],    "ofe":[33,33,50,44],           "may":[0,0,0,0],            "upcs":[1053,1040,1041,1038],"cats":[21,21,21,21]},
  {"c":"LA NACIONAL",        "tot":[12559,12817,12950,11401],"nor":[12474,12817,12908,11389],"ofe":[116,0,64,12],           "may":[453,541,369,389],    "upcs":[706,711,723,646],    "cats":[89,91,93,85]},
  {"c":"LA TERMINAL",        "tot":[3136,3024,2991,2723],    "nor":[3136,3024,2989,2723],    "ofe":[10,0,15,6],             "may":[99,110,108,78],      "upcs":[425,411,407,368],    "cats":[81,82,79,77]},
  {"c":"MEGA SUPER",         "tot":[92981,92653,94114,90448],"nor":[92848,92211,93919,89796],"ofe":[4592,5373,4254,6436],   "may":[18279,10569,5621,6542],"upcs":[1459,1462,1482,1428],"cats":[119,120,121,118]},
  {"c":"MEGASUPER PULPOU",   "tot":[19817,19936,19929,19894],"nor":[19817,19936,19929,19894],"ofe":[5339,5395,5407,5391],   "may":[0,0,17,0],           "upcs":[2823,2830,2829,2824],"cats":[123,123,123,123]},
  {"c":"PERIMERCADOS",       "tot":[0,11235,12768,11249],    "nor":[0,11235,12768,11249],    "ofe":[0,0,0,0],               "may":[0,0,0,0],            "upcs":[0,1602,1821,1604],   "cats":[0,122,124,124]},
  {"c":"PRICESMART",         "tot":[1631,1631,1610,1610],    "nor":[1631,1631,1610,1610],    "ofe":[24,28,0,21],            "may":[0,0,0,0],            "upcs":[216,216,214,214],    "cats":[75,75,75,75]},
  {"c":"SANTA CECI MARKET",  "tot":[5838,5747,6027,5705],    "nor":[5838,5747,6027,5705],    "ofe":[0,0,6,6],               "may":[96,59,72,84],        "upcs":[807,794,835,790],    "cats":[95,98,102,97]},
  {"c":"SUPER COMPRO",       "tot":[26788,27167,26460,26182],"nor":[26782,27118,26460,26026],"ofe":[69,108,0,254],          "may":[42,31,0,49],         "upcs":[2251,2259,2223,2201],"cats":[136,135,135,136]},
  {"c":"SUPER DRAGON",       "tot":[2254,2247,2380,1932],    "nor":[2254,2247,2380,1932],    "ofe":[0,0,12,0],              "may":[48,106,90,72],       "upcs":[279,273,289,243],    "cats":[77,76,76,75]},
  {"c":"SUPER EL DOLAR",     "tot":[2772,2380,903,1288],     "nor":[2772,2380,903,1288],     "ofe":[0,0,0,0],               "may":[12,28,0,0],          "upcs":[387,334,128,180],    "cats":[100,92,28,50]},
  {"c":"SUPER PAGUE MENOS",  "tot":[2799,2355,742,792],      "nor":[2751,2212,742,728],      "ofe":[192,151,0,64],          "may":[3,19,0,0],           "upcs":[402,348,106,119],    "cats":[100,92,27,28]},
 ],
 "GT": [
  {"c":"CRUZ VERDE",         "tot":[4508,4501,4459,4459],    "nor":[4508,4501,4459,4459],    "ofe":[152,98,42,36],          "may":[0,0,0,0],            "upcs":[640,639,633,633],    "cats":[18,18,18,18]},
  {"c":"DOLLARCITY",         "tot":[1288,1309,1232,1239],    "nor":[1288,1309,1232,1239],    "ofe":[0,0,0,0],               "may":[48,40,32,48],        "upcs":[168,169,158,158],    "cats":[51,52,52,52]},
  {"c":"FARMA VALUE",        "tot":[3654,3640,3668,3612],    "nor":[3654,3640,3668,3612],    "ofe":[0,0,0,0],               "may":[0,0,0,0],            "upcs":[518,516,520,512],    "cats":[17,17,17,17]},
  {"c":"GALENO",             "tot":[3934,3927,3885,3885],    "nor":[3934,3927,3885,3885],    "ofe":[18,30,30,24],           "may":[0,0,0,0],            "upcs":[558,557,551,551],    "cats":[18,18,18,18]},
  {"c":"LA BODEGONA",        "tot":[7719,8249,8234,7855],    "nor":[7686,8211,8190,7819],    "ofe":[1112,1368,1349,1238],   "may":[330,465,345,326],    "upcs":[950,1023,1020,969],  "cats":[123,126,125,125]},
  {"c":"LA TORRE",           "tot":[28729,29779,30310,28344],"nor":[28399,29162,29960,28098],"ofe":[5037,5921,5584,5296],   "may":[1220,1727,1960,1182],"upcs":[1420,1412,1401,1393],"cats":[137,136,136,137]},
  {"c":"LA TORRE PULPOU",    "tot":[30359,30219,29554,29428],"nor":[30359,30219,29554,29428],"ofe":[1200,422,431,152],      "may":[0,0,0,0],            "upcs":[3848,3828,3734,3717],"cats":[144,144,143,143]},
  {"c":"MI SUPER FRESH",     "tot":[8183,7721,7700,7712],    "nor":[8183,7714,7700,7700],    "ofe":[767,762,606,550],       "may":[211,204,174,150],    "upcs":[1051,990,987,989],   "cats":[122,121,121,123]},
  {"c":"PRICESMART",         "tot":[1638,1638,1652,1659],    "nor":[1638,1638,1652,1659],    "ofe":[54,63,0,11],            "may":[0,0,0,0],            "upcs":[212,212,214,215],    "cats":[70,70,72,72]},
  {"c":"SUMA",               "tot":[20122,20120,20036,19964],"nor":[20118,20104,20020,19957],"ofe":[2237,2671,2892,2086],   "may":[12184,12636,11268,11800],"upcs":[1271,1272,1266,1261],"cats":[128,129,128,126]},
  {"c":"SUPER 24",           "tot":[1113,1169,1281,1211],    "nor":[1113,1169,1281,1211],    "ofe":[0,84,0,0],              "may":[0,4,0,0],            "upcs":[63,61,65,57],        "cats":[8,9,9,8]},
  {"c":"SUPER DEL BARRIO",   "tot":[12414,12674,12626,12488],"nor":[12404,12656,12600,12488],"ofe":[1088,1756,1702,1026],   "may":[234,370,313,220],    "upcs":[739,757,752,743],    "cats":[110,111,110,109]},
 ],
 "HN": [
  {"c":"BANASUPRO",          "tot":[119,119,119,119],        "nor":[119,119,119,119],        "ofe":[0,0,0,0],               "may":[0,0,0,0],            "upcs":[14,14,14,14],        "cats":[12,12,12,12]},
  {"c":"EL COLONIAL",        "tot":[7140,7125,7098,7154],    "nor":[7140,7119,7098,7154],    "ofe":[503,450,210,365],       "may":[0,0,0,0],            "upcs":[947,945,941,949],    "cats":[120,120,121,120]},
  {"c":"LA COLONIA",         "tot":[31780,31302,31787,31433],"nor":[31780,31290,31787,31430],"ofe":[1697,1836,1498,966],    "may":[235,321,302,231],    "upcs":[1339,1331,1334,1325],"cats":[128,127,127,127]},
  {"c":"LA COLONIA PULPOU",  "tot":[17542,17920,18221,18256],"nor":[17542,17920,18221,18256],"ofe":[0,926,0,373],           "may":[0,0,0,0],            "upcs":[2309,2361,2402,2407],"cats":[138,143,146,145]},
  {"c":"PRICESMART",         "tot":[1659,1652,1645,1659],    "nor":[1659,1652,1645,1659],    "ofe":[48,56,6,36],            "may":[0,0,0,0],            "upcs":[222,221,220,222],    "cats":[71,71,71,71]},
 ],
 "NI": [
  {"c":"AHORRA MAS",         "tot":[6678,6748,6797,7091],    "nor":[6678,6748,6797,7091],    "ofe":[896,991,662,701],       "may":[212,275,172,189],    "upcs":[934,944,951,991],    "cats":[124,124,124,129]},
  {"c":"CALDERA",            "tot":[35,42,49,35],            "nor":[35,42,49,35],            "ofe":[0,0,0,0],               "may":[0,0,0,0],            "upcs":[5,6,7,5],            "cats":[1,1,1,1]},
  {"c":"FARMA VALUE",        "tot":[1792,1778,1785,1799],    "nor":[1792,1778,1785,1799],    "ofe":[0,0,0,0],               "may":[0,0,0,0],            "upcs":[255,253,254,256],    "cats":[14,14,14,14]},
  {"c":"LA COLONIA",         "tot":[6930,7665,7791,7959],    "nor":[6930,7665,7791,7959],    "ofe":[700,728,607,770],       "may":[218,198,177,231],    "upcs":[966,1071,1089,1113], "cats":[124,124,126,129]},
  {"c":"MERCADO ORIENTAL",   "tot":[854,917,980,924],        "nor":[854,917,980,924],        "ofe":[0,0,0,0],               "may":[0,0,0,0],            "upcs":[120,129,138,130],    "cats":[37,38,40,38]},
  {"c":"PRICESMART",         "tot":[1092,1057,1078,1085],    "nor":[1092,1057,1078,1085],    "ofe":[24,28,0,7],             "may":[0,0,0,0],            "upcs":[154,150,152,153],    "cats":[60,59,59,59]},
  {"c":"SAN MARTIN",         "tot":[147,154,154,161],        "nor":[147,154,154,161],        "ofe":[0,0,0,0],               "may":[0,0,0,0],            "upcs":[21,22,22,23],        "cats":[2,2,2,2]},
  {"c":"SUPER EL SELECTO",   "tot":[2205,2226,2226,2408],    "nor":[2205,2226,2226,2408],    "ofe":[0,56,0,60],             "may":[0,6,0,3],            "upcs":[301,304,304,332],    "cats":[88,91,91,95]},
  {"c":"SUPER EXPRESS",      "tot":[448,448,455,539],        "nor":[448,448,455,539],        "ofe":[35,35,28,19],           "may":[6,6,0,0],            "upcs":[64,64,65,77],        "cats":[34,34,34,43]},
  {"c":"SUPER LA ESTRELLA",  "tot":[2324,2324,2583,2324],    "nor":[2324,2324,2583,2324],    "ofe":[98,0,140,0],            "may":[16,0,71,0],          "upcs":[321,321,356,321],    "cats":[91,91,99,91]},
  {"c":"SUPER LA SEGOVIA",   "tot":[2324,2324,2583,2324],    "nor":[2324,2324,2583,2324],    "ofe":[28,0,7,0],              "may":[6,0,6,0],            "upcs":[318,318,355,318],    "cats":[89,89,98,89]},
 ],
 "SV": [
  {"c":"DOLLARCITY",                  "tot":[1428,1379,1414,1463],    "nor":[1428,1379,1414,1463],    "ofe":[0,0,0,0],           "may":[0,0,0,0],        "upcs":[200,193,198,205],    "cats":[57,57,58,58]},
  {"c":"EL BARATILLO",                "tot":[1687,1974,2058,2149],    "nor":[1687,1974,2058,2149],    "ofe":[0,0,21,21],         "may":[0,0,0,0],        "upcs":[240,276,288,300],    "cats":[77,82,80,85]},
  {"c":"FARMA VALUE",                 "tot":[2198,2177,2149,2156],    "nor":[2198,2177,2149,2156],    "ofe":[1709,1630,1416,1602],"may":[0,0,0,0],        "upcs":[313,310,306,307],    "cats":[19,19,18,18]},
  {"c":"FARMACIA SAN NICOLAS",        "tot":[2219,2191,2142,2149],    "nor":[2219,2191,2142,2149],    "ofe":[1716,1692,1086,1692],"may":[0,0,0,0],        "upcs":[317,313,306,307],    "cats":[18,18,17,17]},
  {"c":"LACTEOS EL RODEO",            "tot":[1589,2002,2002,2289],    "nor":[1589,2002,2002,2289],    "ofe":[0,63,0,72],         "may":[0,0,0,0],        "upcs":[227,280,280,321],    "cats":[75,81,81,86]},
  {"c":"PRICESMART",                  "tot":[1190,1190,1148,1190],    "nor":[1190,1190,1148,1190],    "ofe":[30,35,0,21],        "may":[0,0,0,0],        "upcs":[168,168,162,168],    "cats":[66,67,67,67]},
  {"c":"SUPER ALAMEDA",               "tot":[1470,1470,2261,1470],    "nor":[1470,1470,2261,1470],    "ofe":[0,0,0,0],           "may":[0,0,0,0],        "upcs":[210,210,321,210],    "cats":[84,84,93,84]},
  {"c":"SUPER MAYOREO NUEVO MILAGRO", "tot":[1862,2233,2387,2499],    "nor":[1862,2233,2387,2499],    "ofe":[6,21,35,54],        "may":[0,0,0,0],        "upcs":[265,314,335,349],    "cats":[79,82,88,92]},
  {"c":"SUPER SELECTOS",              "tot":[42770,42651,42531,42525],"nor":[42770,42651,42525,42525],"ofe":[5786,5615,4985,5727],"may":[1364,1511,1399,1573],"upcs":[2994,2986,2978,2977],"cats":[140,141,141,141]},
 ],
}

PAIS_NAMES = {
    "CR": "Costa Rica",
    "GT": "Guatemala",
    "HN": "Honduras",
    "NI": "Nicaragua",
    "SV": "El Salvador",
}

# ──────────────────────────────────────────────────────────────────
#  HELPERS DE ESTILO
# ──────────────────────────────────────────────────────────────────
BLUE_DARK  = "0053E2"
BLUE_MED   = "003AB0"
YELLOW_WMT = "FFC220"
GREEN_WMT  = "2A8703"
RED_WMT    = "EA1100"
WHITE      = "FFFFFF"
GRAY_LIGHT = "F0F4FF"
GRAY_SUBTLE= "F5F5F5"
GRAY_BDR   = "C4C4C4"

def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def font(bold=False, color="2E2E2E", size=10):
    return Font(bold=bold, color=color, size=size)

def border():
    s = Side(style="thin", color=GRAY_BDR)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center")

def safe_pct(a, b):
    return round((a - b) / b * 100, 2) if b and b != 0 else None

# Colores de formato condicional
FILL_GREEN  = fill("C8E6C9"); FONT_GREEN  = Font(bold=True, color="1B5E20", size=9)
FILL_RED    = fill("FFCDD2"); FONT_RED    = Font(bold=True, color="B71C1C", size=9)
FILL_YELLOW = fill("FFF9C4"); FONT_YELLOW = Font(bold=True, color="795548", size=9)

def apply_delta_cf(ws, col_letters, row_start, row_end):
    """Aplica formato condicional a columnas de delta."""
    for cl in col_letters:
        rng = f"{cl}{row_start}:{cl}{row_end}"
        ws.conditional_formatting.add(rng, CellIsRule("greaterThan", ["0"],  fill=FILL_GREEN,  font=FONT_GREEN))
        ws.conditional_formatting.add(rng, CellIsRule("lessThan",    ["0"],  fill=FILL_RED,    font=FONT_RED))
        ws.conditional_formatting.add(rng, CellIsRule("equal",       ["0"],  fill=FILL_YELLOW, font=FONT_YELLOW))

def write_cell(ws, row, col, value, bg=WHITE, bold=False, txt_color="2E2E2E",
               align="center", num_fmt=None, sz=9):
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = fill(bg)
    c.font   = Font(bold=bold, color=txt_color, size=sz)
    c.alignment = center() if align == "center" else left()
    c.border = border()
    if num_fmt:
        c.number_format = num_fmt
    return c

# ──────────────────────────────────────────────────────────────────
#  WORKBOOK
# ──────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════
#  HOJAS POR PAIS — Detalle Semanal
# ══════════════════════════════════════════════════════════════════
for pais_code, rows in DATA.items():

    ws = wb.create_sheet(title=pais_code)

    # — Titulo —
    last_col = get_column_letter(27)
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value = (f"Detalle Semanal — Todos los Competidores | {PAIS_NAMES[pais_code]} | "
               f"SW11–SW14 · 2026 | Divisiones excluidas: MG, TEXTIL")
    t.fill = fill(BLUE_DARK); t.font = Font(bold=True, color=WHITE, size=12)
    t.alignment = center(); ws.row_dimensions[1].height = 28

    # — Fila 2: grupos de SW —
    groups = [
        (2,  4,  "SW11",                 BLUE_DARK),
        (5,  7,  "SW12",                 BLUE_MED),
        (8,  10, "SW13",                 BLUE_MED),
        (11, 13, "SW14",                 BLUE_DARK),
        (14, 16, "Delta Absoluto 14/13", "7C0000"),
        (17, 19, "Delta %  14/13",       "A03000"),
        (20, 22, "Delta Absoluto 14/12", "7C4500"),
        (23, 25, "Delta %  14/12",       "9B5E00"),
    ]
    for sc, ec, label, color in groups:
        sl = get_column_letter(sc); el = get_column_letter(ec)
        ws.merge_cells(f"{sl}2:{el}2")
        c = ws[f"{sl}2"]
        c.value = label; c.fill = fill(color)
        c.font = Font(bold=True, color=WHITE, size=9)
        c.alignment = center(); c.border = border()

    for col, label in [(26, "UPCs SW14"), (27, "Cats SW14")]:
        c = ws.cell(row=2, column=col, value=label)
        c.fill = fill("767676"); c.font = Font(bold=True, color=WHITE, size=9)
        c.alignment = center(); c.border = border()
    ws.row_dimensions[2].height = 24

    # — Fila 3: sub-encabezados —
    sub = ["Competidor",
           "Normal","Oferta","Mayoreo",  # SW11
           "Normal","Oferta","Mayoreo",  # SW12
           "Normal","Oferta","Mayoreo",  # SW13
           "Normal","Oferta","Mayoreo",  # SW14
           "D Nor","D Ofe","D May",      # delta abs 14/13
           "D% Nor","D% Ofe","D% May",  # delta %  14/13
           "D Nor","D Ofe","D May",      # delta abs 14/12
           "D% Nor","D% Ofe","D% May",  # delta %  14/12
           "UPCs","Cats"]
    for ci, h in enumerate(sub, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = fill(YELLOW_WMT); c.font = Font(bold=True, color="2E2E2E", size=9)
        c.alignment = center(); c.border = border()
    ws.row_dimensions[3].height = 22

    # — Datos —
    pct_cols = [get_column_letter(c) for c in [17,18,19,23,24,25]]
    abs_cols = [get_column_letter(c) for c in [14,15,16,20,21,22]]

    for ri, r in enumerate(rows, start=4):
        bg = WHITE if ri % 2 == 0 else GRAY_LIGHT

        # Competidor
        write_cell(ws, ri, 1, r["c"], bg=bg, bold=True, txt_color=BLUE_DARK, align="left")

        # Datos SW11-SW14
        for i, sw_fields in enumerate(zip(r["nor"], r["ofe"], r["may"])):
            base_col = 2 + i * 3
            for j, val in enumerate(sw_fields):
                write_cell(ws, ri, base_col+j, val, bg=bg, num_fmt="#,##0")

        # Delta absoluto 14/13
        for offset, field in enumerate(["nor","ofe","may"]):
            d = r[field][3] - r[field][2]
            write_cell(ws, ri, 14+offset, d, bg=bg, num_fmt="+#,##0;-#,##0;0")

        # Delta % 14/13
        for offset, field in enumerate(["nor","ofe","may"]):
            p = safe_pct(r[field][3], r[field][2])
            if p is not None:
                write_cell(ws, ri, 17+offset, p/100, bg=bg, num_fmt="+0.0%;-0.0%;0.0%")
            else:
                write_cell(ws, ri, 17+offset, None, bg=bg)

        # Delta absoluto 14/12
        for offset, field in enumerate(["nor","ofe","may"]):
            d = r[field][3] - r[field][1]
            write_cell(ws, ri, 20+offset, d, bg=bg, num_fmt="+#,##0;-#,##0;0")

        # Delta % 14/12
        for offset, field in enumerate(["nor","ofe","may"]):
            p = safe_pct(r[field][3], r[field][1])
            if p is not None:
                write_cell(ws, ri, 23+offset, p/100, bg=bg, num_fmt="+0.0%;-0.0%;0.0%")
            else:
                write_cell(ws, ri, 23+offset, None, bg=bg)

        # UPCs y Cats
        write_cell(ws, ri, 26, r["upcs"][3], bg=bg, num_fmt="#,##0")
        write_cell(ws, ri, 27, r["cats"][3], bg=bg, num_fmt="#,##0")

    data_end = 3 + len(rows)
    apply_delta_cf(ws, abs_cols + pct_cols, 4, data_end)

    # — Anchos —
    ws.column_dimensions["A"].width = 28
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 11
    for col in range(14, 26):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.column_dimensions[get_column_letter(26)].width = 9
    ws.column_dimensions[get_column_letter(27)].width = 7

    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A3:{get_column_letter(27)}{data_end}"


# ══════════════════════════════════════════════════════════════════
#  HOJA TENDENCIAS
# ══════════════════════════════════════════════════════════════════
def tendencia_sostenida(tot):
    """True si la serie es monotona (solo sube o solo baja)."""
    diffs = [tot[i+1] - tot[i] for i in range(len(tot)-1)]
    return all(d >= 0 for d in diffs) or all(d <= 0 for d in diffs)

# Calcular clasificacion
rising, falling = [], []
for pais_code, rows in DATA.items():
    for r in rows:
        t11, t14 = r["tot"][0], r["tot"][3]
        if t11 == 0:
            continue  # PERIMERCADOS sin base, no clasificable
        tot_pct  = safe_pct(t14, t11)
        ofe_abs  = r["ofe"][3] - r["ofe"][0]
        ofe_pct  = safe_pct(r["ofe"][3], r["ofe"][0]) if r["ofe"][0] > 0 else None
        may_abs  = r["may"][3] - r["may"][0]
        may_pct  = safe_pct(r["may"][3], r["may"][0]) if r["may"][0] > 0 else None
        nor_abs  = r["nor"][3] - r["nor"][0]
        nor_pct  = safe_pct(r["nor"][3], r["nor"][0]) if r["nor"][0] > 0 else None
        sostenida = tendencia_sostenida(r["tot"])
        obj = {
            "pais": pais_code, "c": r["c"],
            "t11": t11, "t14": t14,
            "tot_abs": t14 - t11, "tot_pct": tot_pct,
            "nor_abs": nor_abs, "nor_pct": nor_pct,
            "ofe_abs": ofe_abs, "ofe_pct": ofe_pct,
            "may_abs": may_abs, "may_pct": may_pct,
            "upcs14": r["upcs"][3], "cats14": r["cats"][3],
            "sostenida": sostenida,
        }
        if tot_pct is not None and tot_pct > 0:
            rising.append(obj)
        else:
            falling.append(obj)

rising.sort(key=lambda x: -(x["tot_pct"] or 0))
falling.sort(key=lambda x: (x["tot_pct"] or 0))

ws_t = wb.create_sheet(title="TENDENCIAS")

# Titulo
ws_t.merge_cells("A1:P1")
tc = ws_t["A1"]
tc.value = "Clasificacion: Tendencias Crecientes vs Decrecientes por Competidor | SW11 vs SW14 · 2026"
tc.fill = fill(BLUE_DARK); tc.font = Font(bold=True, color=WHITE, size=12)
tc.alignment = center(); ws_t.row_dimensions[1].height = 28

# Encabezados
HEADERS_T = [
    "Clasificacion", "Pais", "Competidor",
    "Tendencia Sostenida",
    "SW11 Total", "SW14 Total",
    "D Total", "D% Total",
    "D Normal", "D% Normal",
    "D Oferta", "D% Oferta",
    "D Mayoreo", "D% Mayoreo",
    "UPCs SW14", "Cats SW14",
]
for ci, h in enumerate(HEADERS_T, 1):
    c = ws_t.cell(row=2, column=ci, value=h)
    c.fill = fill(YELLOW_WMT); c.font = Font(bold=True, color="2E2E2E", size=9)
    c.alignment = center(); c.border = border()
ws_t.row_dimensions[2].height = 22

PCT_FMT = "+0.00%;-0.00%;0.00%"
NUM_FMT = "+#,##0;-#,##0;0"

def write_trend_row(ws, row_idx, group, obj):
    # Colores por grupo
    cls_bg  = "E8F5E2" if group == "CRECIENTE" else "FFEBEE"
    cls_fc  = "1B5E20" if group == "CRECIENTE" else "B71C1C"
    row_bg  = "F1FBF1" if group == "CRECIENTE" else "FFF5F5"
    if row_idx % 2 == 0:
        row_bg = "FFFFFF"

    write_cell(ws, row_idx, 1, group,           bg=cls_bg, bold=True, txt_color=cls_fc)
    write_cell(ws, row_idx, 2, obj["pais"],     bg=row_bg, bold=True, txt_color=BLUE_DARK)
    write_cell(ws, row_idx, 3, obj["c"],        bg=row_bg, bold=True, txt_color="2E2E2E", align="left")
    write_cell(ws, row_idx, 4, "Si" if obj["sostenida"] else "No", bg=row_bg)
    write_cell(ws, row_idx, 5, obj["t11"],      bg=row_bg, num_fmt="#,##0")
    write_cell(ws, row_idx, 6, obj["t14"],      bg=row_bg, num_fmt="#,##0")
    write_cell(ws, row_idx, 7, obj["tot_abs"],  bg=row_bg, num_fmt=NUM_FMT)
    p = obj["tot_pct"]
    write_cell(ws, row_idx, 8, p/100 if p is not None else None, bg=row_bg, num_fmt=PCT_FMT)

    for offset, key_abs, key_pct in [
        (0, "nor_abs", "nor_pct"),
        (2, "ofe_abs", "ofe_pct"),
        (4, "may_abs", "may_pct"),
    ]:
        write_cell(ws, row_idx, 9+offset,  obj[key_abs], bg=row_bg, num_fmt=NUM_FMT)
        pp = obj[key_pct]
        write_cell(ws, row_idx, 10+offset, pp/100 if pp is not None else None, bg=row_bg, num_fmt=PCT_FMT)

    write_cell(ws, row_idx, 15, obj["upcs14"],  bg=row_bg, num_fmt="#,##0")
    write_cell(ws, row_idx, 16, obj["cats14"],  bg=row_bg, num_fmt="#,##0")

row_idx = 3
for obj in rising:
    write_trend_row(ws_t, row_idx, "CRECIENTE",   obj); row_idx += 1
for obj in falling:
    write_trend_row(ws_t, row_idx, "DECRECIENTE", obj); row_idx += 1

data_end_t = row_idx - 1

# Formato condicional en columnas de delta (7-14)
delta_cols_t = [get_column_letter(c) for c in range(7, 15)]
apply_delta_cf(ws_t, delta_cols_t, 3, data_end_t)

# Formato condicional columna 1 (CRECIENTE/DECRECIENTE)
rng_cls = f"A3:A{data_end_t}"
ws_t.conditional_formatting.add(rng_cls, CellIsRule("equal", ['"CRECIENTE"'],
    fill=fill("C8E6C9"), font=Font(bold=True, color="1B5E20", size=9)))
ws_t.conditional_formatting.add(rng_cls, CellIsRule("equal", ['"DECRECIENTE"'],
    fill=fill("FFCDD2"), font=Font(bold=True, color="B71C1C", size=9)))

# Anchos
ws_t.column_dimensions["A"].width = 14
ws_t.column_dimensions["B"].width = 7
ws_t.column_dimensions["C"].width = 28
ws_t.column_dimensions["D"].width = 12
for col in range(5, 17):
    ws_t.column_dimensions[get_column_letter(col)].width = 11

ws_t.freeze_panes = "D3"
ws_t.auto_filter.ref = f"A2:P{data_end_t}"


# ══════════════════════════════════════════════════════════════════
#  HOJA RESUMEN
# ══════════════════════════════════════════════════════════════════
ws_res = wb.create_sheet(title="RESUMEN", index=0)
ws_res.merge_cells("A1:G1")
t = ws_res["A1"]
t.value = "Resumen General — Todos los Paises · SW11–SW14 · 2026"
t.fill = fill(BLUE_DARK); t.font = Font(bold=True, color=WHITE, size=13)
t.alignment = center(); ws_res.row_dimensions[1].height = 32

for ci, h in enumerate(["Pais","Competidores","Total Normal SW14","Total Ofertas SW14",
                         "Total Mayoreo SW14","UPCs Prom SW14","Tendencia General"], 1):
    c = ws_res.cell(row=2, column=ci, value=h)
    c.fill = fill(YELLOW_WMT); c.font = Font(bold=True, color="2E2E2E", size=10)
    c.alignment = center(); c.border = border()
ws_res.row_dimensions[2].height = 20

for ri, (code, rows) in enumerate(DATA.items(), start=3):
    bg = WHITE if ri % 2 == 0 else GRAY_SUBTLE
    total_nor = sum(r["nor"][3] for r in rows)
    total_ofe = sum(r["ofe"][3] for r in rows)
    total_may = sum(r["may"][3] for r in rows)
    avg_upcs  = round(sum(r["upcs"][3] for r in rows) / len(rows))
    total_sw11= sum(r["tot"][0] for r in rows if r["tot"][0] > 0)
    total_sw14= sum(r["tot"][3] for r in rows)
    trend_pct = safe_pct(total_sw14, total_sw11)
    trend_lbl = (f"+{trend_pct:.1f}%" if trend_pct and trend_pct > 0 else
                 f"{trend_pct:.1f}%" if trend_pct else "N/A")
    trend_col = "1B5E20" if trend_pct and trend_pct > 0 else "B71C1C"

    for ci, (val, fmt, aln) in enumerate([
        (f"{PAIS_NAMES[code]} ({code})", None, "left"),
        (len(rows), "#,##0", "center"),
        (total_nor, "#,##0", "center"),
        (total_ofe, "#,##0", "center"),
        (total_may, "#,##0", "center"),
        (avg_upcs,  "#,##0", "center"),
        (trend_lbl, None, "center"),
    ], 1):
        c = ws_res.cell(row=ri, column=ci, value=val)
        c.fill = fill(bg); c.border = border()
        c.font = Font(bold=(ci==1), color=BLUE_DARK if ci==1 else (trend_col if ci==7 else "2E2E2E"), size=10)
        c.alignment = center() if aln == "center" else left()
        if fmt: c.number_format = fmt

for col, w in zip("ABCDEFG", [28,12,16,16,16,13,14]):
    ws_res.column_dimensions[col].width = w


# ──────────────────────────────────────────────────────────────────
#  GUARDAR
# ──────────────────────────────────────────────────────────────────
today = date.today().strftime("%Y%m%d")
out_path = rf"C:\Users\shern22\Documents\puppy_workspace\Detalle_Semanal_ConTendencias_{today}.xlsx"
wb.save(out_path)
print(f"OK:{out_path}")
