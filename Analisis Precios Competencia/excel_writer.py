"""
excel_writer.py  —  Genera el Excel actualizado SW12-SW15.
"""
import math
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Paleta Walmart ────────────────────────────────────────────────────────────
BLUE   = "0053E2"
DKBLUE = "002B7A"
SPARK  = "FFC220"
GREEN  = "2A8703"
RED    = "EA1100"
WHITE  = "FFFFFF"
LGRAY  = "F5F5F5"
MGRAY  = "E0E0E0"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=WHITE, size=10):
    return Font(bold=bold, color=color, size=size)

def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border():
    thin = Side(style="thin", color=MGRAY)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

COUNTRY_NAMES = {
    "CR": "Costa Rica",
    "GT": "Guatemala",
    "HN": "Honduras",
    "NI": "Nicaragua",
    "SV": "El Salvador",
}

WEEKS = ["SW12", "SW13", "SW14", "SW15"]
WEEK_KEYS = ["sw12", "sw13", "sw14", "sw15"]
PRICE_TYPES = ["Normal", "Oferta", "Mayoreo"]
PRICE_KEYS  = ["nor", "ofe", "may"]


def _write_country_sheet(ws, code, rows):
    country_name = COUNTRY_NAMES[code]

    # ── Fila 1: título ────────────────────────────────────────────────────────
    title = (
        f"Detalle Semanal – Todos los Competidores | {country_name} | "
        f"SW12–SW15 – 2026 | Divisiones excluidas: MG, TEXTIL"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=27)
    c = ws.cell(1, 1, title)
    c.font = _font(bold=True, color=WHITE, size=11)
    c.fill = _fill(DKBLUE)
    c.alignment = _align("left")
    ws.row_dimensions[1].height = 22

    # ── Fila 2: grupos de semanas + deltas + UPCs ─────────────────────────────
    groups = [
        ("SW12", 1, 3), ("SW13", 4, 6), ("SW14", 7, 9), ("SW15", 10, 12),
        ("Delta Absoluto 15/14", 13, 15), ("Delta %  15/14", 16, 18),
        ("Delta Absoluto 15/13", 19, 21), ("Delta %  15/13", 22, 24),
    ]
    for lbl, c_start, c_end in groups:
        col_hex = BLUE if "SW" in lbl else DKBLUE
        ws.merge_cells(
            start_row=2, start_column=c_start + 1,
            end_row=2,   end_column=c_end + 1
        )
        cell = ws.cell(2, c_start + 1, lbl)
        cell.fill = _fill(col_hex)
        cell.font = _font(bold=True)
        cell.alignment = _align()

    # UPCs / Cats SW15
    for col, lbl in [(26, "UPCs SW15"), (27, "Cats SW15")]:
        c2 = ws.cell(2, col, lbl)
        c2.fill = _fill(DKBLUE)
        c2.font = _font(bold=True)
        c2.alignment = _align()

    ws.row_dimensions[2].height = 18

    # ── Fila 3: sub-headers ───────────────────────────────────────────────────
    sub_headers = ["Competidor"]
    for _ in WEEKS:
        sub_headers += ["Normal", "Oferta", "Mayoreo"]
    # delta 15/14
    sub_headers += ["D Nor", "D Ofe", "D May", "D% Nor", "D% Ofe", "D% May"]
    # delta 15/13
    sub_headers += ["D Nor", "D Ofe", "D May", "D% Nor", "D% Ofe", "D% May"]
    sub_headers += ["UPCs", "Cats"]

    for j, lbl in enumerate(sub_headers, start=1):
        cell = ws.cell(3, j, lbl)
        cell.fill = _fill(BLUE)
        cell.font = _font(bold=False, size=9)
        cell.alignment = _align()
    ws.row_dimensions[3].height = 14

    # ── Datos ─────────────────────────────────────────────────────────────────
    for r_idx, row in enumerate(rows, start=4):
        ws.row_dimensions[r_idx].height = 13
        fill = _fill(LGRAY) if r_idx % 2 == 0 else None
        col = 1

        # Competidor
        c_cell = ws.cell(r_idx, col, row["competidor"])
        c_cell.font = Font(bold=True, size=9)
        c_cell.alignment = _align("left", wrap=False)
        if fill:
            c_cell.fill = fill
        col += 1

        # Semanas
        for wk in WEEK_KEYS:
            for pt in PRICE_KEYS:
                val = row.get(f"{wk}_{pt}")
                nc = ws.cell(r_idx, col, val if val is not None else "")
                nc.alignment = _align()
                nc.font = Font(size=9)
                if fill:
                    nc.fill = fill
                if val is not None:
                    nc.number_format = "#,##0"
                col += 1

        # Deltas absolutos 15/14
        for pt in PRICE_KEYS:
            val = row.get(f"d1514_{pt}")
            nc = ws.cell(r_idx, col, val if val is not None else "")
            nc.alignment = _align()
            nc.font = Font(
                size=9,
                color=GREEN if val and float(val) > 0 else (RED if val and float(val) < 0 else "000000")
            )
            if fill:
                nc.fill = fill
            if val is not None:
                nc.number_format = "#,##0"
            col += 1

        # Deltas % 15/14
        for pt in PRICE_KEYS:
            val = row.get(f"d1514_{pt}_pct")
            nc = ws.cell(r_idx, col, val if val is not None else "")
            nc.alignment = _align()
            nc.font = Font(
                size=9,
                color=GREEN if val and float(val) > 0 else (RED if val and float(val) < 0 else "000000")
            )
            if fill:
                nc.fill = fill
            if val is not None:
                nc.number_format = "0.00%"
            col += 1

        # Deltas absolutos 15/13
        for pt in PRICE_KEYS:
            val = row.get(f"d1513_{pt}")
            nc = ws.cell(r_idx, col, val if val is not None else "")
            nc.alignment = _align()
            nc.font = Font(
                size=9,
                color=GREEN if val and float(val) > 0 else (RED if val and float(val) < 0 else "000000")
            )
            if fill:
                nc.fill = fill
            if val is not None:
                nc.number_format = "#,##0"
            col += 1

        # Deltas % 15/13
        for pt in PRICE_KEYS:
            val = row.get(f"d1513_{pt}_pct")
            nc = ws.cell(r_idx, col, val if val is not None else "")
            nc.alignment = _align()
            nc.font = Font(
                size=9,
                color=GREEN if val and float(val) > 0 else (RED if val and float(val) < 0 else "000000")
            )
            if fill:
                nc.fill = fill
            if val is not None:
                nc.number_format = "0.00%"
            col += 1

        # UPCs / Cats
        for key in ["upcs_sw15", "cats_sw15"]:
            val = row.get(key)
            nc = ws.cell(r_idx, col, val if val is not None else "")
            nc.alignment = _align()
            nc.font = Font(size=9)
            if fill:
                nc.fill = fill
            if val is not None:
                nc.number_format = "#,##0"
            col += 1

    # ── Anchos de columna ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 26
    for col_i in range(2, 28):
        ws.column_dimensions[get_column_letter(col_i)].width = 10
    ws.freeze_panes = "B4"


def _write_resumen(ws, resumen_rows):
    ws.merge_cells("A1:G1")
    c = ws.cell(1, 1, "Resumen General – Todos los Países – SW12–SW15 – 2026")
    c.font = _font(bold=True, size=12)
    c.fill = _fill(DKBLUE)
    c.alignment = _align("left")
    ws.row_dimensions[1].height = 24

    headers = [
        "País", "Competidores",
        "Total Normal SW15", "Total Ofertas SW15",
        "Total Mayoreo SW15", "UPCs Prom SW15", "Tendencia General"
    ]
    for j, h in enumerate(headers, 1):
        c2 = ws.cell(2, j, h)
        c2.fill = _fill(BLUE)
        c2.font = _font(bold=True)
        c2.alignment = _align()
    ws.row_dimensions[2].height = 16

    for i, row in enumerate(resumen_rows, 3):
        fill = _fill(LGRAY) if i % 2 == 0 else None
        vals = [
            row["pais"], row["competidores"],
            row["total_normal_sw15"], row["total_ofertas_sw15"],
            row["total_mayoreo_sw15"], row["upcs_prom_sw15"],
            row["tendencia_general"],
        ]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(i, j, v)
            cell.alignment = _align("left" if j == 1 else "center")
            cell.font = Font(size=10, bold=(j == 7))
            if fill:
                cell.fill = fill
            if j in (3, 4, 5, 6) and v:
                cell.number_format = "#,##0"

    col_ws = [26, 14, 16, 16, 16, 14, 14]
    for i, w in enumerate(col_ws, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_tendencias(ws, tend_rows):
    ws.merge_cells("A1:P1")
    c = ws.cell(1, 1,
        "Clasificacion: Tendencias Crecientes vs Decrecientes por "
        "Competidor | SW12 vs SW15 – 2026"
    )
    c.font = _font(bold=True, size=11)
    c.fill = _fill(DKBLUE)
    c.alignment = _align("left")
    ws.row_dimensions[1].height = 22

    headers = [
        "Clasificacion", "Pais", "Competidor", "Tendencia Sostenida",
        "SW12 Total", "SW15 Total",
        "D Total", "D% Total",
        "D Normal", "D% Normal",
        "D Oferta", "D% Oferta",
        "D Mayoreo", "D% Mayoreo",
        "UPCs SW15", "Cats SW15",
    ]
    for j, h in enumerate(headers, 1):
        c2 = ws.cell(2, j, h)
        c2.fill = _fill(BLUE)
        c2.font = _font(bold=True, size=9)
        c2.alignment = _align()
    ws.row_dimensions[2].height = 16

    for i, row in enumerate(tend_rows, 3):
        is_creciente = row["clasificacion"] == "CRECIENTE"
        fill = _fill("E8F5E9") if is_creciente else _fill("FFEBEE")
        base_color = GREEN if is_creciente else RED

        vals = [
            row["clasificacion"], row["pais"], row["competidor"],
            row["tendencia_sostenida"],
            row["sw12_total"], row["sw15_total"],
            row["d_total"], row["d_pct_total"],
            row["d_normal"], row["d_pct_normal"],
            row["d_oferta"], row["d_pct_oferta"],
            row["d_mayoreo"], row["d_pct_mayoreo"],
            row["upcs_sw15"], row["cats_sw15"],
        ]
        pct_cols = {8, 10, 12, 14}  # D% columns (1-indexed)
        for j, v in enumerate(vals, 1):
            cell = ws.cell(i, j, v if v is not None else "")
            cell.fill = fill
            cell.alignment = _align(
                "left" if j in (1, 2, 3) else "center", wrap=False
            )
            cell.font = Font(
                size=9,
                bold=(j == 1),
                color=base_color if j == 1 else "000000"
            )
            if j in pct_cols and v:
                cell.number_format = "0.00%"
            elif j in (5, 6, 7, 9, 11, 13, 15, 16) and v:
                cell.number_format = "#,##0"

    col_widths = [14, 6, 28, 18, 12, 12, 10, 9, 10, 9, 10, 9, 10, 9, 10, 9]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "E3"


def write_excel(data, out_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quita la hoja default

    # RESUMEN
    ws_res = wb.create_sheet("RESUMEN")
    _write_resumen(ws_res, data["resumen"])

    # Países
    for code, rows in data["countries"].items():
        ws = wb.create_sheet(code)
        _write_country_sheet(ws, code, rows)

    # TENDENCIAS
    ws_tend = wb.create_sheet("TENDENCIAS")
    _write_tendencias(ws_tend, data["tendencias"])

    wb.save(out_path)
    print(f"[OK] Excel guardado: {out_path}", flush=True)
