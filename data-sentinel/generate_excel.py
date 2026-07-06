"""
DataSentinel – Generador de catálogo Excel
Ejecutar: uv run generate_excel.py
"""

import os
from datetime import date


from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Colores Walmart ───────────────────────────────────────────────
BLUE       = "FF0053E2"   # wblue.100
BLUE_LIGHT = "FFD6E4FF"
SPARK      = "FFFFC220"   # spark.100
WHITE      = "FFFFFFFF"
GRAY_H     = "FFF4F6FA"   # fila par
GRAY_B     = "FFE8EDF5"   # borde header

# Sensibilidad
GREEN_BG   = "FFD4EDDA"; GREEN_FG   = "FF155724"
LBLUE_BG   = "FFD1ECF1"; LBLUE_FG   = "FF0C5460"
AMBER_BG   = "FFFFF3CD"; AMBER_FG   = "FF856404"
RED_BG     = "FFF8D7DA"; RED_FG     = "FF721C24"

# Estado
GREN2_BG   = "FFD4EDDA"; GREN2_FG   = "FF155724"
YELL_BG    = "FFFFF3CD"; YELL_FG    = "FF856404"
CYAN_BG    = "FFD1ECF1"; CYAN_FG    = "FF0C5460"
REDD_BG    = "FFF8D7DA"; REDD_FG    = "FF721C24"

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(hex_, bold=False, sz=10):
    return Font(color=hex_, bold=bold, size=sz, name="Segoe UI")
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def thin_border():
    s = Side(style="thin", color="FFD0D7E3")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Columnas ──────────────────────────────────────────────────────
COLUMNS = [
    ("ID",           8),
    ("Nombre del Activo",  38),
    ("Tipo",         16),
    ("Plataforma",   14),
    ("Descripción",  46),
    ("Link / Ubicación",  36),
    ("Responsable",  22),
    ("Área / Depto", 22),
    ("Sensibilidad", 16),
    ("Estado",       16),
    ("PII",           8),
    ("Financiero",    10),
    ("Notas",        40),
    ("Fecha Registro", 14),
]

# ── Datos de muestra ──────────────────────────────────────────────
SAMPLE_DATA = [
    (1, "walmart_mx.analytics.ventas_diarias", "Tabla BigQuery", "BigQuery",
     "Ventas por tienda, categoría y SKU. Actualización diaria vía POS + OMS.",
     "", "Carlos Mendoza", "Comercial",
     "Confidencial", "Validado", "No", "Sí",
     "Acceso: grupo bq-ventas-comercial. Retención 3 años.", "01/06/2025"),

    (2, "Dashboard KPIs Semanales", "Dashboard", "Looker",
     "KPIs de sell-through, margen y tráfico por formato de tienda.",
     "https://looker.walmart.com/dashboards/kpi-semanal",
     "Ana García", "Data & Analytics",
     "Interna", "Validado", "No", "No",
     "Visible para todos los empleados con cuenta Walmart.", "01/06/2025"),

    (3, "Reporte RH – Nómina y Ausencias", "Reporte", "Power BI",
     "Nómina, absentismo y evaluaciones de desempeño del personal activo.",
     "", "Patricia Ruiz", "Recursos Humanos",
     "Restringida", "En revisión", "Sí", "Sí",
     "⚠️ Pendiente encriptación a nivel columna. Deadline: 31-ene-2025.", "01/06/2025"),

    (4, "walmart_mx.supply.proveedores_master", "Tabla BigQuery", "BigQuery",
     "Maestro de proveedores: contactos, contratos y evaluaciones de desempeño.",
     "", "Roberto López", "Compras",
     "Interna", "Validado", "No", "No",
     "Retención 7 años (cumplimiento fiscal).", "01/06/2025"),

    (5, "Ops Métricas – Tráfico y Conversión", "Dashboard", "Tableau",
     "Tráfico, conversión, tiempos de espera en caja e inventario fuera de stock.",
     "https://tableau.walmart.com/views/ops-metricas",
     "José Hernández", "Operaciones",
     "Interna", "Pendiente", "No", "No",
     "Migración a Looker en Q3-2025.", "01/06/2025"),

    (6, "walmart_mx.rrhh.nomina_dataset", "Tabla BigQuery", "BigQuery",
     "Dataset de nómina con salarios, deducciones y prestaciones.",
     "", "Silvia Torres", "Finanzas",
     "Restringida", "No cumple", "Sí", "Sí",
     "CRÍTICO: encriptación a nivel columna no implementada.", "01/06/2025"),
]

# ── Validaciones ──────────────────────────────────────────────────
DV_TIPO = '"Tabla BigQuery,Dashboard,Reporte,Vista (View),API,Archivo,Otro"'
DV_PLAT = '"BigQuery,Looker,Power BI,Tableau,SharePoint,GCS,Hive,Teradata,Otro"'
DV_SENS = '"Pública,Interna,Confidencial,Restringida"'
DV_ESTD = '"Validado,Pendiente,En revisión,No cumple"'
DV_BOOL = '"Sí,No"'

def make_dv(formula, col, rows=200):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.sqref = f"{col}3:{col}{rows}"
    return dv

def build_excel(path: str):
    wb = Workbook()

    # ── Hoja principal ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Catálogo"
    ws.sheet_view.showGridLines = False

    # Dimensiones
    for i, (_, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32

    # ── Fila 1: título corporativo ────────────────────────────────
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = "🛡️  DataSentinel — Catálogo de Gobernanza de Datos"
    c.fill = fill(BLUE)
    c.font = font(WHITE, bold=True, sz=13)
    c.alignment = center()
    c.border = thin_border()

    # ── Fila 2: encabezados ───────────────────────────────────────
    for i, (label, _) in enumerate(COLUMNS, 1):
        c = ws.cell(row=2, column=i, value=label)
        c.fill = fill("FF1A3A6B")
        c.font = font(WHITE, bold=True, sz=10)
        c.alignment = center()
        c.border = thin_border()

    # ── Datos ─────────────────────────────────────────────────────
    for row_idx, row in enumerate(SAMPLE_DATA, 3):
        bg = GRAY_H if row_idx % 2 == 0 else WHITE
        for col_idx, val in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.fill = fill(bg)
            c.font = font("FF1F2937", sz=10)
            c.alignment = left() if col_idx in (2, 5, 6, 13) else center()
            c.border = thin_border()

    # ── Filas vacías para nuevos registros ────────────────────────
    for row_idx in range(len(SAMPLE_DATA) + 3, 203):
        bg = GRAY_H if row_idx % 2 == 0 else WHITE
        for col_idx in range(1, len(COLUMNS) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.fill = fill(bg)
            c.font = font("FF1F2937", sz=10)
            c.alignment = left() if col_idx in (2, 5, 6, 13) else center()
            c.border = thin_border()

    # ── Validación de datos (dropdowns) ──────────────────────────
    col_map = {3: DV_TIPO, 4: DV_PLAT, 9: DV_SENS, 10: DV_ESTD, 11: DV_BOOL, 12: DV_BOOL}
    for col_num, formula in col_map.items():
        ws.add_data_validation(make_dv(formula, get_column_letter(col_num)))

    # ── Formato condicional – Sensibilidad (col I = 9) ────────────
    col_s = "I"
    rules_s = [
        ("Pública",      GREEN_BG, GREEN_FG),
        ("Interna",      LBLUE_BG, LBLUE_FG),
        ("Confidencial", AMBER_BG, AMBER_FG),
        ("Restringida",  RED_BG,   RED_FG),
    ]
    for val, bg, fg in rules_s:
        ws.conditional_formatting.add(
            f"{col_s}3:{col_s}202",
            FormulaRule(
                formula=[f'${col_s}3="{val}"'],
                fill=fill(bg[2:]),
                font=Font(color=fg[2:], bold=True, size=10, name="Segoe UI"),
            )
        )

    # ── Formato condicional – Estado (col J = 10) ─────────────────
    col_e = "J"
    rules_e = [
        ("Validado",    GREN2_BG, GREN2_FG),
        ("Pendiente",   YELL_BG,  YELL_FG),
        ("En revisión", CYAN_BG,  CYAN_FG),
        ("No cumple",   REDD_BG,  REDD_FG),
    ]
    for val, bg, fg in rules_e:
        ws.conditional_formatting.add(
            f"{col_e}3:{col_e}202",
            FormulaRule(
                formula=[f'${col_e}3="{val}"'],
                fill=fill(bg[2:]),
                font=Font(color=fg[2:], bold=True, size=10, name="Segoe UI"),
            )
        )

    # ── Tabla Excel (permite filtros por columna) ─────────────────
    last_row = len(SAMPLE_DATA) + 2
    table = Table(displayName="Catalogo", ref=f"A2:N{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)

    # ── Freeze panes (fijar primeras 2 filas) ─────────────────────
    ws.freeze_panes = "A3"

    # ── Hoja de referencia ────────────────────────────────────────
    ws2 = wb.create_sheet("📖 Guía de Campos")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 40

    guide = [
        ("Campo", "Descripción", "Valores aceptados"),
        ("Nombre del Activo", "Ruta técnica o nombre del recurso de datos",
         "ej. project.dataset.table / Nombre Dashboard"),
        ("Tipo", "Categoría del activo de datos",
         "Tabla BigQuery, Dashboard, Reporte, Vista, API, Archivo, Otro"),
        ("Plataforma", "Sistema donde reside el activo",
         "BigQuery, Looker, Power BI, Tableau, SharePoint…"),
        ("Descripción", "Qué contiene, para qué se usa, qué áreas lo consumen", "Texto libre"),
        ("Link / Ubicación", "URL directa o path técnico completo", "URL o ruta"),
        ("Responsable", "Nombre del Data Owner del activo", "Nombre completo"),
        ("Área / Depto", "Equipo dueño del activo", "ej. Data & Analytics, RH, Finanzas…"),
        ("Sensibilidad",
         "Pública: sin restricciones | Interna: solo empleados | "
         "Confidencial: acceso limitado | Restringida: máxima protección",
         "Pública / Interna / Confidencial / Restringida"),
        ("Estado", "Nivel de estandarización y cumplimiento del activo",
         "Validado / Pendiente / En revisión / No cumple"),
        ("PII", "¿Contiene datos personales identificables?", "Sí / No"),
        ("Financiero", "¿Contiene datos financieros sensibles?", "Sí / No"),
        ("Notas", "Restricciones de acceso, política de retención, alertas", "Texto libre"),
        ("Fecha Registro", "Fecha en que se registró el activo", "DD/MM/AAAA"),
    ]

    ws2.row_dimensions[1].height = 28
    for r, (campo, desc, vals) in enumerate(guide, 1):
        is_header = r == 1
        for c, val in enumerate([campo, desc, vals], 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.fill = fill("FF1A3A6B" if is_header else (GRAY_H if r % 2 == 0 else WHITE))
            cell.font = font(WHITE if is_header else "FF1F2937", bold=is_header, sz=10)
            cell.alignment = left()
            cell.border = thin_border()

    # ── Guardar ───────────────────────────────────────────────────
    wb.save(path)
    print(f"OK  Excel generado: {path}")

if __name__ == "__main__":
    out = os.path.join(os.path.dirname(__file__), "DataSentinel_Catalogo.xlsx")
    build_excel(out)
    os.startfile(out)
